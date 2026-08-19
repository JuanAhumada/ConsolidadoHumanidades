"""Escritura del Excel de salida (estilos, colores de fila, anchos)."""
from __future__ import annotations

import re
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from consolidado.config.settings import construir_columnas_salida, construir_grupos_encabezado, etiqueta_export_columna
from consolidado.core.columnas import alinear_dataframe_salida, formatear_dataframe_salida
from consolidado.core.constants import (
    ANCHO_MAXIMO_COLUMNA_EXCEL,
    COL_ACTIVACION_RUTA,
    COL_ACTIVOS,
    COL_ALERTA_PROPIA,
    COL_AJUSTE_RAZONABLE,
    COL_FECHA_NACIMIENTO,
    COL_NUM_ALERTA_FINAL,
    COL_NUM_ALERTA_INICIAL,
    COL_PERIODO_ACTUAL,
    COL_TELEFONO_CELULAR,
    COL_TIPO_ALERTA_FINAL,
    COL_TIPO_ALERTA_INICIAL,
    COL_TOTAL_BECA,
    FONT_MATERIA_REPETIDA,
    FORMATO_FECHA_DMY,
    HOJA_LISTADO,
    _cfg,
    max_materias_en_dataframe,
)
from consolidado.core.excel_io import _longitud_visible_celda
from consolidado.core.normalizacion import (
    _es_nulo,
    _es_valor_true,
    es_estudiante_activo,
    formatear_fecha_nacimiento,
    formatear_periodo_cod,
    normalizar_id,
    normalizar_telefono_celda,
)
from consolidado.core.prioridad import color_excel_fila
from consolidado.core.repetidas import _materia_es_repetida

FONT_TITULO_GRUPO_EXCEL = Font(bold=True, size=14)
FONT_ENCABEZADO_COLUMNA_EXCEL = Font(bold=True, size=11)
FONT_DATOS_EXCEL = Font(size=11)

_BORDE_DELGADO = Side(style="thin", color="C0C0C0")
_BORDE_GRUESO = Side(style="medium", color="555555")


def _borde_celda(
    col: int,
    fila: int,
    ultima_col: int,
    inicios_grupo: set[int],
) -> Border:
    izquierda = _BORDE_DELGADO
    derecha = _BORDE_DELGADO
    if col in inicios_grupo and col > 1:
        izquierda = _BORDE_GRUESO
    arriba = _BORDE_DELGADO if fila > 1 else None
    abajo = _BORDE_DELGADO
    return Border(left=izquierda, right=derecha, top=arriba, bottom=abajo)


def _aplicar_bordes_y_fuentes_encabezado(
    ws,
    *,
    ultima_fila: int,
    ultima_col: int,
    inicios_grupo: set[int],
) -> None:
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    for fila in range(1, ultima_fila + 1):
        for col in range(1, ultima_col + 1):
            celda = ws.cell(row=fila, column=col)
            celda.border = _borde_celda(col, fila, ultima_col, inicios_grupo)
            if fila == 1:
                celda.font = FONT_TITULO_GRUPO_EXCEL
                celda.alignment = Alignment(horizontal="center", vertical="center")
            elif fila == 2:
                celda.font = FONT_ENCABEZADO_COLUMNA_EXCEL
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif fila >= 3:
                if not (celda.font and celda.font.underline == "single" and celda.font.bold):
                    celda.font = FONT_DATOS_EXCEL
                celda.alignment = Alignment(vertical="top", wrap_text=True)


def _valor_excel_celda(col: str, val):
    if _es_nulo(val):
        return None
    if col == "Priorizado":
        return True if val is True or _es_valor_true(val) else None
    if col == COL_ACTIVOS:
        return es_estudiante_activo(val)
    if col == COL_TELEFONO_CELULAR:
        norm = normalizar_telefono_celda(val)
        if not norm:
            return None
        partes = norm.split(", ")
        if len(partes) == 1:
            try:
                return int(partes[0])
            except ValueError:
                return partes[0]
        return norm
    if col == "Identificación":
        if isinstance(val, int):
            return val
        digits = re.sub(r"\D", "", str(val))
        if digits:
            try:
                return int(digits)
            except ValueError:
                pass
    if col == "Periodo ingreso" and isinstance(val, (int, float)) and not _es_nulo(val):
        if isinstance(val, float) and val == int(val):
            return int(val)
    if col == COL_PERIODO_ACTUAL:
        return formatear_periodo_cod(val) or val
    if col == COL_FECHA_NACIMIENTO:
        return formatear_fecha_nacimiento(val, FORMATO_FECHA_DMY) or val
    if col in (COL_NUM_ALERTA_INICIAL, COL_NUM_ALERTA_FINAL) and not _es_nulo(val):
        try:
            return int(float(str(val).strip()))
        except ValueError:
            return val
    if col == COL_TOTAL_BECA and not _es_nulo(val):
        try:
            n = float(str(val).strip().replace(",", ""))
            return int(n) if n == int(n) else n
        except ValueError:
            return val
    if col in (COL_ACTIVACION_RUTA, COL_ALERTA_PROPIA):
        if val is True or _es_valor_true(val):
            return True
        return None
    return val


def _relleno_fila(row: dict) -> PatternFill | None:
    codigo = color_excel_fila(row)
    if not codigo:
        return None
    return PatternFill(start_color=codigo, end_color=codigo, fill_type="solid")


def _ajustar_hoja(ws, num_filas: int, num_cols: int, *, fila_encabezado: int = 2) -> None:
    if num_filas < fila_encabezado or num_cols < 1:
        return

    ultima_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{fila_encabezado}:{ultima_col}{num_filas}"
    ws.freeze_panes = f"A{fila_encabezado + 1}"
    alineacion_envuelta = Alignment(wrap_text=True, vertical="top")

    for idx in range(1, num_cols + 1):
        letter = get_column_letter(idx)
        max_len = 0
        for row in range(1, num_filas + 1):
            celda = ws.cell(row, idx)
            v = celda.value
            if v is not None:
                max_len = max(max_len, _longitud_visible_celda(v))
            if row > fila_encabezado and max_len > 80:
                celda.alignment = alineacion_envuelta
        ancho = min(max(max_len + 2, 10), ANCHO_MAXIMO_COLUMNA_EXCEL)
        ws.column_dimensions[letter].width = ancho

def _escribir_hoja_consolidada(
    ws,
    df: pl.DataFrame,
    columnas: list[str],
    grupos: list[tuple[str, list[str]]],
    *,
    materias_repetidas: dict[str, set[str]] | None = None,
) -> int:
    """Fila 1: encabezados principales unificados. Fila 2: columnas. Fila 3+: datos."""
    materias_repetidas = materias_repetidas or {}
    datos = formatear_dataframe_salida(alinear_dataframe_salida(df, columnas))
    col_idx = 1
    mapa_pos: dict[str, int] = {}
    inicios_grupo: set[int] = set()

    for nombre_grupo, cols_grupo in grupos:
        if not cols_grupo:
            continue
        inicio = col_idx
        inicios_grupo.add(inicio)
        for col in cols_grupo:
            if col not in columnas:
                continue
            mapa_pos[col] = col_idx
            ws.cell(row=2, column=col_idx, value=etiqueta_export_columna(col))
            col_idx += 1
        fin = col_idx - 1
        if inicio <= fin:
            ws.merge_cells(start_row=1, start_column=inicio, end_row=1, end_column=fin)
            celda_grupo = ws.cell(row=1, column=inicio, value=nombre_grupo)
            celda_grupo.alignment = Alignment(horizontal="center", vertical="center")
            celda_grupo.font = FONT_TITULO_GRUPO_EXCEL

    ultima_fila = 2
    for r_idx, row in enumerate(datos.iter_rows(named=True), start=3):
        ultima_fila = r_idx
        relleno_fila = _relleno_fila(row)
        id_key = normalizar_id(row.get("Identificación"))
        repetidas_est = materias_repetidas.get(id_key, set()) if id_key else set()
        for col, c_idx in mapa_pos.items():
            celda = ws.cell(
                row=r_idx,
                column=c_idx,
                value=_valor_excel_celda(col, row.get(col)),
            )
            if relleno_fila is not None:
                celda.fill = relleno_fila
            if repetidas_est and col.startswith("Materia ") and _materia_es_repetida(
                row.get(col), repetidas_est
            ):
                celda.font = FONT_MATERIA_REPETIDA

    ultima_col = len(mapa_pos)
    _aplicar_bordes_y_fuentes_encabezado(
        ws,
        ultima_fila=ultima_fila,
        ultima_col=ultima_col,
        inicios_grupo=inicios_grupo,
    )
    _ajustar_hoja(ws, ultima_fila, ultima_col, fila_encabezado=2)
    return ultima_fila

def _guardar_workbook_excel(wb: Workbook, ruta: Path) -> Path:
    """
    Guarda el libro. Si la ruta está bloqueada (p. ej. abierta en Excel),
    guarda una copia con marca de tiempo en el mismo directorio.
    """
    destino = ruta.resolve()
    destino.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(destino)
        wb.close()
        return destino
    except PermissionError as exc:
        raise PermissionError(
            f"No se puede guardar el consolidado en:\n  {destino}\n\n"
            "Cierra el archivo de salida si lo tienes abierto en Excel "
            "(o cualquier programa que lo use) y vuelve a intentar."
        ) from exc

def guardar_excel_consolidado(
    consolidado: pl.DataFrame,
    ruta: Path,
    *,
    cfg: dict | None = None,
    num_materias: int | None = None,
    materias_repetidas: dict[str, set[str]] | None = None,
) -> Path:
    """Una sola hoja con encabezados principales (Datos, Priorizados, Becas, Materias)."""
    cfg = cfg or _cfg()
    n_mat = num_materias if num_materias is not None else max_materias_en_dataframe(consolidado)
    columnas = construir_columnas_salida(cfg, n_mat)
    grupos = construir_grupos_encabezado(cfg, n_mat)
    nombre_hoja = cfg.get("salida", {}).get("hoja", HOJA_LISTADO)

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    _escribir_hoja_consolidada(
        ws, consolidado, columnas, grupos, materias_repetidas=materias_repetidas
    )
    return _guardar_workbook_excel(wb, ruta)

