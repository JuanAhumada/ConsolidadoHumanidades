from __future__ import annotations

import subprocess
import sys
import time
import tkinter as tk
from pathlib import Path

from consolidado.paths import PROJECT_ROOT
from tkinter import filedialog, messagebox
from zipfile import BadZipFile

import polars as pl
from openpyxl import load_workbook

from consolidado.config.settings import guardar_config
from consolidado.core.constants import (
    ERRORES_LECTURA_REINTENTABLES,
    MAX_REINTENTOS_LECTURA,
    PAUSA_ENTRE_REINTENTOS_SEG,
    TITULOS_EXCEL_FUENTE,
    _cfg,
)
def _longitud_visible_celda(valor) -> int:
    if valor is None:
        return 0
    texto = str(valor).replace("\r\n", "\n")
    return max((len(linea) for linea in texto.split("\n")), default=0)

def abrir_archivo_en_sistema(ruta: Path, *, parent: tk.Misc | None = None) -> None:
    """Abre el archivo generado con la aplicación predeterminada (Excel en Windows)."""
    destino = ruta.resolve()
    if not destino.is_file():
        return

    def _abrir() -> None:
        try:
            if sys.platform == "win32":
                # os.startfile puede provocar fatal error PyEval_RestoreThread
                # al llamarse desde callbacks de Tkinter en Python 3.13+.
                subprocess.Popen(
                    ["cmd", "/c", "start", "", str(destino)],
                    close_fds=True,
                )
            elif sys.platform == "darwin":
                subprocess.run(["open", str(destino)], check=False)
            else:
                subprocess.run(["xdg-open", str(destino)], check=False)
        except OSError as exc:
            print(f"No se pudo abrir automáticamente el archivo: {exc}")

    if parent is not None:
        try:
            parent.after(1, _abrir)
            return
        except tk.TclError:
            pass
    _abrir()

def _pedir_nueva_ruta_excel(etiqueta: str, ruta_fallida: Path) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass
    root.update_idletasks()
    try:
        nueva = filedialog.askopenfilename(
            title=f"{etiqueta} — seleccione el archivo de nuevo",
            filetypes=[
                ("Libro Excel", "*.xlsx *.xlsm *.xls"),
                ("Todos los archivos", "*.*"),
            ],
            initialdir=str(ruta_fallida.parent) if ruta_fallida.parent.exists() else None,
            initialfile=ruta_fallida.name,
        )
    finally:
        root.destroy()
    if not nueva:
        return None
    p = Path(nueva)
    return p if p.is_file() else None

def _dialogo_fallo_lectura(ruta: Path, etiqueta: str, error: Exception) -> str:
    """'reintentar' | 'elegir' | 'cancelar'."""
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass
    root.update_idletasks()
    mensaje = (
        f"No se pudo leer el archivo ({etiqueta}):\n\n"
        f"{ruta}\n\n"
        f"Error: {error}\n\n"
        "Cierre el archivo en Excel si está abierto.\n\n"
        "Sí = Reintentar el mismo archivo\n"
        "No = Elegir otro archivo\n"
        "Cancelar = Detener el proceso"
    )
    try:
        respuesta = messagebox.askyesnocancel("Error al leer Excel", mensaje, icon="warning")
    finally:
        root.destroy()
    if respuesta is True:
        return "reintentar"
    if respuesta is False:
        return "elegir"
    return "cancelar"

def ejecutar_lectura_con_recuperacion(
    ruta: Path,
    etiqueta: str,
    operacion,
    *,
    permitir_seleccionar_otro: bool = True,
):
    """
    Ejecuta operacion(ruta). Reintenta lectura; si falla, pregunta reintentar o elegir otro archivo.
    Devuelve (resultado, ruta_utilizada).
    """
    ruta_actual = Path(ruta)
    ultimo_error: Exception | None = None

    while True:
        for intento in range(1, MAX_REINTENTOS_LECTURA + 1):
            try:
                return operacion(ruta_actual), ruta_actual
            except ERRORES_LECTURA_REINTENTABLES as exc:
                ultimo_error = exc
                if intento < MAX_REINTENTOS_LECTURA:
                    time.sleep(PAUSA_ENTRE_REINTENTOS_SEG)
                    continue
                break
            except Exception as exc:
                ultimo_error = exc
                break

        if not permitir_seleccionar_otro:
            raise ultimo_error  # type: ignore[misc]

        accion = _dialogo_fallo_lectura(ruta_actual, etiqueta, ultimo_error)  # type: ignore[arg-type]
        if accion == "reintentar":
            continue
        if accion == "elegir":
            nueva = _pedir_nueva_ruta_excel(etiqueta, ruta_actual)
            if nueva is None:
                raise SystemExit("Operación cancelada: no se seleccionó un archivo de reemplazo.")
            ruta_actual = nueva
            continue
        raise SystemExit(f"Operación cancelada al leer: {ruta_actual}")

def _nombres_hojas_excel(ruta: Path) -> list[str]:
    suf = ruta.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        wb = load_workbook(ruta, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    if suf == ".xls":
        return ["Sheet1"]
    raise ValueError(f"Formato no soportado: {ruta.name} (usa .xlsx, .xlsm o .xls)")

def _leer_hoja_excel(ruta: Path, hoja: str) -> pl.DataFrame:
    suf = ruta.suffix.lower()
    if suf in (".xlsx", ".xlsm", ".xls"):
        return pl.read_excel(ruta, sheet_name=hoja, infer_schema_length=0)
    raise ValueError(f"Formato no soportado: {ruta.name} (usa .xlsx, .xlsm o .xls)")

def elegir_cuatro_excels_en_explorador() -> list[Path] | None:
    """Cuatro diálogos 'Abrir archivo'. Devuelve None si el usuario cancela."""
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass
    root.update_idletasks()

    tipos = [
        ("Libro Excel", "*.xlsx *.xlsm *.xls"),
        ("Todos los archivos", "*.*"),
    ]
    elegidos: list[Path] = []
    try:
        for etiqueta in TITULOS_EXCEL_FUENTE:
            while True:
                ruta = filedialog.askopenfilename(
                    title=f"{etiqueta} — elige este Excel (solo lectura; no se mueve el archivo)",
                    filetypes=tipos,
                )
                if not ruta:
                    return None
                p = Path(ruta)
                if not p.is_file():
                    messagebox.showerror("Error", f"No existe el archivo:\n{p}")
                    continue
                if p in elegidos:
                    messagebox.showwarning(
                        "Archivo repetido",
                        "Ese archivo ya fue seleccionado. Elige otro distinto.",
                    )
                    continue
                elegidos.append(p)
                break
    finally:
        root.destroy()

    return elegidos

def resolver_ruta_salida_consolidado(
    cfg: dict | None = None,
    base: Path | None = None,
    *,
    parent=None,
) -> Path | None:
    """Pregunta si sobrescribir el último consolidado; devuelve None si se cancela."""
    base = base or PROJECT_ROOT
    cfg = cfg or _cfg()
    rel_salida = cfg.get("salida", {}).get("ruta", "salida/estudiantes_consolidado.xlsx")
    destino = (base / rel_salida).resolve()

    if destino.is_file():
        resp = messagebox.askyesnocancel(
            "Sobrescribir consolidado",
            f"Ya existe un consolidado en:\n  {destino}\n\n"
            "¿Desea sobrescribirlo?\n\n"
            "Sí = sobrescribir\n"
            "No = elegir otra ruta\n"
            "Cancelar = abortar",
            parent=parent,
        )
        if resp is None:
            return None
        if resp is False:
            elegida = elegir_guardar_consolidado(destino)
            if elegida is None:
                return None
            destino = elegida.resolve()
            try:
                rel_nueva = destino.relative_to(base.resolve())
                cfg.setdefault("salida", {})["ruta"] = rel_nueva.as_posix()
            except ValueError:
                cfg.setdefault("salida", {})["ruta"] = str(destino)
            guardar_config(cfg, base)

    destino.parent.mkdir(parents=True, exist_ok=True)
    return destino

def elegir_guardar_consolidado(ruta_por_defecto: Path) -> Path | None:
    """Diálogo Guardar como. None = cancelar (no escribir)."""
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass
    root.update_idletasks()
    try:
        ruta_por_defecto.parent.mkdir(parents=True, exist_ok=True)
        inicial_dir = str(ruta_por_defecto.parent)
    except OSError:
        inicial_dir = str(Path.home())

    try:
        guardar = filedialog.asksaveasfilename(
            title="Guardar Excel consolidado",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfile=ruta_por_defecto.name,
            initialdir=inicial_dir,
        )
    finally:
        root.destroy()

    if not guardar:
        return None
    return Path(guardar)

def _es_archivo_excel_valido(p: Path) -> bool:
    return p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")

def listar_excels_en_carpeta(entrada: Path) -> list[Path]:
    archivos = sorted(entrada.glob("*.xlsx")) + sorted(entrada.glob("*.xlsm"))
    return [p for p in archivos if _es_archivo_excel_valido(p)]

