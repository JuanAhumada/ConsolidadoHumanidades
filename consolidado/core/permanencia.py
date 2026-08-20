"""
Permanencia y ruta de grado.

Cruza estudiantes por documento (cohortes + base general) y lee las hojas de
metas. El left join no crea filas nuevas: solo enriquece al consolidado.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

from consolidado.config.settings import COLUMNAS_RUTA_GRADO, carpeta_excels
from consolidado.core.constants import COL_ACTIVOS
from consolidado.core.normalizacion import (
    _es_nulo,
    formatear_periodo_cod,
    normalizar_encabezado,
    normalizar_id,
    programa_es_permitido,
)

_VACIOS = frozenset({"", "-", "—", "–", ".", "n/a", "na", "none", "nan", "null"})
_ENCABEZADOS_META = frozenset({"programas", "programa"})
_FILAS_TITULO_META = frozenset(
    {
        "programas",
        "programa",
        "proyeccion",
        "proyeccion estimada",
        "graduacion",
        "permanencia",
    }
)

_ALIAS_CAMPOS: dict[str, tuple[str, ...]] = {
    "pct_creditos": (
        "% creditos aprobados",
        "% aprobados",
        "% creditos aprobados.",
    ),
    "estado_opcion": ("estado de opcion de grado",),
    "opcion": ("opcion de grado", "opcion a grado"),
    "estado_ingles": ("estado de ingles",),
    "saber_pro": ("saber pro2", "saber pro"),
    "observaciones": (
        "observacion de seguimiento",
        "observaciones de seguimiento",
        "observaciones",
        "observacion",
    ),
    "periodo_grado": ("periodo grado", "periodo de grado"),
    "estado_graduacion": (
        "estado a 20261",
        "estado a 20262",
        "estado a 20251",
        "estado a 20252",
        "estado",
    ),
}

_SALIDA = {
    "pct_creditos": COLUMNAS_RUTA_GRADO[0],
    "estado_opcion": COLUMNAS_RUTA_GRADO[1],
    "opcion": COLUMNAS_RUTA_GRADO[2],
    "estado_ingles": COLUMNAS_RUTA_GRADO[3],
    "saber_pro": COLUMNAS_RUTA_GRADO[4],
    "periodo_grado": "Periodo grado",
    "estado_graduacion": "Estado graduación",
    "cohorte_graduacion": "Cohorte de graduación",
}

_RE_HOJA_COHORTE = re.compile(
    r"cohorte\s+(\d{4})\s*[-–]\s*([12])(?:\s*\(\s*(\d{4})\s*[-–]\s*([12])\s*\))?",
    re.IGNORECASE,
)
_SEMESTRES_CARRERA = 12

_ALIAS_META = {
    "programa": ("programas", "programa"),
    "periodo_inicio": ("periodo inicio", "periodo de inicio"),
    "poblacion": ("poblacion",),
    "meta_num": ("meta #", "meta n", "meta num"),
    "alcanzado_num": ("# alcanzado", "alcanzado"),
    "faltante_num": ("# faltante", "faltante"),
    "meta_pct": ("meta %",),
    "alcanzado_pct": ("% alcanzado",),
    "faltante_pct": ("% faltante",),
    "estado": ("estado",),
}


def _texto(val: Any) -> str:
    if _es_nulo(val):
        return ""
    return " ".join(str(val).replace("\xa0", " ").replace("\n", " ").split())


def _es_graduado_obs(val: Any) -> bool:
    return "graduado" in _norm(val)


def _es_vacio(val: Any) -> bool:
    t = _texto(val)
    return not t or t.lower() in _VACIOS


def _norm(val: Any) -> str:
    return normalizar_encabezado(_texto(val))


def _periodo_o_vacio(val: Any) -> str | None:
    p = formatear_periodo_cod(val)
    if p:
        return p
    t = _texto_campo(val)
    if not t or t in {"0", "0.0"}:
        return None
    return t


def _clave_periodo(periodo: str) -> int | None:
    p = formatear_periodo_cod(periodo) or (
        periodo if re.match(r"^\d{4}-[12]$", str(periodo or "")) else None
    )
    if not p:
        return None
    try:
        anio, sem = int(p[:4]), int(p[-1])
    except ValueError:
        return None
    return anio * 2 + (sem - 1)


def _periodo_desde_clave(clave: int) -> str:
    return f"{clave // 2}-{1 if clave % 2 == 0 else 2}"


def _sumar_semestres(periodo: str, n: int) -> str | None:
    clave = _clave_periodo(periodo)
    if clave is None:
        return None
    return _periodo_desde_clave(clave + n)


def _periodo_es_anterior(a: str | None, b: str | None) -> bool:
    ka, kb = _clave_periodo(a or ""), _clave_periodo(b or "")
    if ka is None or kb is None:
        return False
    return ka < kb


def _cohorte_esperada(nombre_hoja: str) -> str | None:
    m = _RE_HOJA_COHORTE.search(_norm(nombre_hoja))
    if not m:
        return None
    ingreso = f"{m.group(1)}-{m.group(2)}"
    paren = f"{m.group(3)}-{m.group(4)}" if m.group(3) else None
    if paren and not _periodo_es_anterior(paren, ingreso):
        return formatear_periodo_cod(paren) or paren
    return _sumar_semestres(ingreso, _SEMESTRES_CARRERA) or paren


def _ruta_slot(cfg: dict, base: Path, carpeta: Path | None, *ids: str) -> Path | None:
    carpeta = Path(carpeta) if carpeta is not None else carpeta_excels(cfg, base)
    buscados = set(ids)
    for slot in cfg.get("archivos_fuente", []):
        if slot.get("tipo") not in buscados and slot.get("id") not in buscados:
            continue
        p = carpeta / slot.get("nombre_guardado", "")
        if p.is_file():
            return p
    return None


def ruta_archivo_permanencia(
    cfg: dict,
    base: Path,
    carpeta: Path | None = None,
) -> Path | None:
    return _ruta_slot(cfg, base, carpeta, "bd_permanencia")


def ruta_archivo_graduacion(
    cfg: dict,
    base: Path,
    carpeta: Path | None = None,
) -> Path | None:
    return _ruta_slot(cfg, base, carpeta, "bd_graduacion")


def _hoja_es_cohorte(nombre: str) -> bool:
    return _norm(nombre).startswith("cohorte")


def _hoja_es_base_general(nombre: str) -> bool:
    n = _norm(nombre)
    return "base" in n and "general" in n


def _hoja_es_metas_graduacion(nombre: str) -> bool:
    n = _norm(nombre)
    return "metas" in n and "gradu" in n


def _hoja_es_metas_permanencia(nombre: str) -> bool:
    n = _norm(nombre)
    return "metas" in n and "perman" in n


def _indices_por_aliases(
    headers_norm: list[str],
    aliases: tuple[str, ...],
    usados: set[int],
) -> list[int]:
    encontrados: list[int] = []
    for alias in aliases:
        for i, h in enumerate(headers_norm):
            if i in usados or h != alias:
                continue
            encontrados.append(i)
            usados.add(i)
    return encontrados


def _indice_por_aliases(
    headers_norm: list[str],
    aliases: tuple[str, ...],
    usados: set[int],
) -> int | None:
    hallados = _indices_por_aliases(headers_norm, aliases, usados)
    return hallados[0] if hallados else None


def _mapa_columnas_estudiante(
    headers: list[Any],
    *,
    fallback_ingles: bool,
) -> dict[str, Any]:
    headers_norm = [_norm(h) for h in headers]
    usados: set[int] = set()
    mapa: dict[str, Any] = {}
    idx_doc = _indice_por_aliases(
        headers_norm, ("documento", "identificacion", "num identificacion"), usados
    )
    if idx_doc is None:
        return {}
    usados.add(idx_doc)
    mapa["documento"] = idx_doc
    for clave, aliases in _ALIAS_CAMPOS.items():
        idxs = _indices_por_aliases(headers_norm, aliases, usados)
        if idxs:
            mapa[clave] = idxs
    if fallback_ingles and "estado_ingles" not in mapa:
        idxs = _indices_por_aliases(headers_norm, ("ingles",), usados)
        if idxs:
            mapa["estado_ingles"] = idxs
    if "estado_graduacion" not in mapa:
        for i, h in enumerate(headers_norm):
            if i in usados:
                continue
            if h.startswith("estado a ") or h == "estado":
                mapa["estado_graduacion"] = [i]
                usados.add(i)
                break
    return mapa


def _es_encabezado_estudiantes(vals: list[Any]) -> bool:
    norms = [_norm(v) for v in vals if not _es_vacio(v)]
    if "documento" not in norms and "identificacion" not in norms:
        return False
    return any("nombre" in n or n == "programa" for n in norms)


def _numero_pct(val: Any) -> str | None:
    if _es_vacio(val):
        return None
    texto = _texto(val).replace("%", "").replace(",", ".")
    try:
        n = float(texto)
    except ValueError:
        return _texto(val)
    if -1.5 <= n <= 1.5:
        n *= 100
    return f"{n:.1f}%".replace(".0%", "%")


def _texto_campo(val: Any) -> str | None:
    if _es_vacio(val):
        return None
    return _texto(val)


def _valor_de_indices(vals: list[Any], indices: list[int], formatear) -> str | None:
    for i in indices:
        if i >= len(vals):
            continue
        valor = formatear(vals[i])
        if valor:
            return valor
    return None


def _fusionar_valores(destino: dict[str, str | None], origen: dict[str, str | None]) -> None:
    """El origen pisa si trae valor (varias hojas de Permanencia)."""
    if origen.get("_graduado"):
        destino["_graduado"] = True
    for clave, valor in origen.items():
        if clave == "_graduado":
            continue
        if valor is None or valor == "":
            continue
        destino[clave] = valor


def _completar_valores(destino: dict[str, str | None], origen: dict[str, str | None]) -> None:
    """El origen solo llena huecos (Gestión + Permanencia)."""
    if origen.get("_graduado"):
        destino["_graduado"] = True
    for clave, valor in origen.items():
        if clave == "_graduado":
            continue
        if valor is None or valor == "":
            continue
        if destino.get(clave):
            continue
        destino[clave] = valor


def _filas_estudiante_hoja(
    ws,
    *,
    fallback_ingles: bool,
    cohorte_graduacion: str | None = None,
) -> dict[str, dict[str, str | None]]:
    out: dict[str, dict[str, str | None]] = {}
    mapa: dict[str, Any] | None = None
    vacias = 0
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if all(_es_vacio(v) for v in vals):
            if mapa is not None:
                vacias += 1
                if vacias >= 40:
                    break
            continue
        vacias = 0
        if _es_encabezado_estudiantes(vals):
            candidato = _mapa_columnas_estudiante(vals, fallback_ingles=fallback_ingles)
            if candidato:
                mapa = candidato
            continue
        if not mapa:
            continue
        idx_doc = mapa["documento"]
        if idx_doc >= len(vals):
            continue
        key = normalizar_id(vals[idx_doc])
        if not key:
            continue
        fila: dict[str, str | None] = {}
        if "pct_creditos" in mapa:
            fila[_SALIDA["pct_creditos"]] = _valor_de_indices(
                vals, mapa["pct_creditos"], _numero_pct
            )
        if "estado_opcion" in mapa:
            fila[_SALIDA["estado_opcion"]] = _valor_de_indices(
                vals, mapa["estado_opcion"], _texto_campo
            )
        if "opcion" in mapa:
            fila[_SALIDA["opcion"]] = _valor_de_indices(vals, mapa["opcion"], _texto_campo)
        if "estado_ingles" in mapa:
            fila[_SALIDA["estado_ingles"]] = _valor_de_indices(
                vals, mapa["estado_ingles"], _texto_campo
            )
        if "saber_pro" in mapa:
            fila[_SALIDA["saber_pro"]] = _valor_de_indices(
                vals, mapa["saber_pro"], _texto_campo
            )
        if "periodo_grado" in mapa:
            fila[_SALIDA["periodo_grado"]] = _valor_de_indices(
                vals, mapa["periodo_grado"], _periodo_o_vacio
            )
        if "estado_graduacion" in mapa:
            estado = _valor_de_indices(vals, mapa["estado_graduacion"], _texto_campo)
            if estado:
                fila[_SALIDA["estado_graduacion"]] = estado
                if "graduado" in _norm(estado):
                    fila["_graduado"] = True
        if "observaciones" in mapa:
            obs = _valor_de_indices(vals, mapa["observaciones"], _texto_campo)
            if _es_graduado_obs(obs):
                fila["_graduado"] = True
        if cohorte_graduacion:
            fila[_SALIDA["cohorte_graduacion"]] = cohorte_graduacion
        previo = out.get(key, {})
        _fusionar_valores(previo, fila)
        out[key] = previo
    return out


def _schema_ruta() -> dict[str, type]:
    return {
        "_id_key": pl.Utf8,
        **{c: pl.Utf8 for c in COLUMNAS_RUTA_GRADO},
        "_graduado": pl.Boolean,
    }


def _df_desde_por_id(por_id: dict[str, dict[str, str | None]]) -> pl.DataFrame:
    schema = _schema_ruta()
    if not por_id:
        return pl.DataFrame(schema=schema)
    filas = []
    for key, data in por_id.items():
        fila: dict[str, Any] = {"_id_key": key}
        for col in COLUMNAS_RUTA_GRADO:
            val = data.get(col)
            fila[col] = str(val) if val is not None else None
        fila["_graduado"] = bool(data.get("_graduado"))
        filas.append(fila)
    return pl.DataFrame(filas, schema=schema)


def leer_estudiantes_permanencia(ruta: Path) -> pl.DataFrame:
    """Devuelve identificación + columnas de ruta de grado, una fila por documento."""
    if not ruta.is_file():
        return pl.DataFrame(schema=_schema_ruta())

    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        por_id: dict[str, dict[str, str | None]] = {}
        hojas_base = [n for n in wb.sheetnames if _hoja_es_base_general(n)]
        hojas_cohorte = [n for n in wb.sheetnames if _hoja_es_cohorte(n)]
        for nombre in hojas_base + hojas_cohorte:
            esperado = _cohorte_esperada(nombre) if _hoja_es_cohorte(nombre) else None
            extra = _filas_estudiante_hoja(
                wb[nombre],
                fallback_ingles=_hoja_es_base_general(nombre),
                cohorte_graduacion=esperado,
            )
            for key, fila in extra.items():
                previo = por_id.get(key, {})
                _fusionar_valores(previo, fila)
                por_id[key] = previo
    finally:
        wb.close()
    return _df_desde_por_id(por_id)


def leer_estudiantes_graduacion(ruta: Path) -> pl.DataFrame:
    """Gestión de graduación: periodo de grado, estado, créditos, saber e inglés."""
    if not ruta.is_file():
        return pl.DataFrame(schema=_schema_ruta())

    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        por_id: dict[str, dict[str, str | None]] = {}
        hojas = [n for n in wb.sheetnames if "gradu" in _norm(n)] or list(wb.sheetnames[:1])
        for nombre in hojas:
            extra = _filas_estudiante_hoja(wb[nombre], fallback_ingles=True)
            for key, fila in extra.items():
                previo = por_id.get(key, {})
                _fusionar_valores(previo, fila)
                por_id[key] = previo
    finally:
        wb.close()
    return _df_desde_por_id(por_id)


def aplicar_permanencia(
    consolidado: pl.DataFrame,
    cfg: dict,
    base: Path,
    carpeta: Path | None = None,
) -> pl.DataFrame:
    """Left join de ruta de grado; no añade estudiantes que no estén en el consolidado."""
    for col in list(COLUMNAS_RUTA_GRADO) + [COL_ACTIVOS]:
        if col in consolidado.columns:
            consolidado = consolidado.drop(col)

    extra = _combinar_graduacion_permanencia(cfg, base, carpeta)

    if extra.height == 0:
        return consolidado.with_columns(
            [pl.lit(None).alias(c) for c in COLUMNAS_RUTA_GRADO]
            + [pl.lit(True).alias(COL_ACTIVOS)]
        )

    resultado = consolidado.with_columns(
        pl.col("Identificación")
        .map_elements(normalizar_id, return_dtype=pl.Utf8)
        .alias("_id_key")
    )
    cols_join = [c for c in extra.columns if c != "_id_key"]
    extra = extra.select(["_id_key"] + cols_join)
    resultado = resultado.join(extra, on="_id_key", how="left").drop("_id_key")
    for col in COLUMNAS_RUTA_GRADO:
        if col not in resultado.columns:
            resultado = resultado.with_columns(pl.lit(None).alias(col))
    if "_graduado" in resultado.columns:
        resultado = resultado.with_columns(
            (~pl.col("_graduado").fill_null(False)).alias(COL_ACTIVOS)
        ).drop("_graduado")
    else:
        resultado = resultado.with_columns(pl.lit(True).alias(COL_ACTIVOS))
    return resultado


def _leer_fuente_ruta(leer, ruta: Path | None) -> pl.DataFrame:
    vacio = pl.DataFrame(schema=_schema_ruta())
    if not ruta:
        return vacio
    try:
        return leer(ruta)
    except Exception:
        return vacio


def _combinar_graduacion_permanencia(
    cfg: dict,
    base: Path,
    carpeta: Path | None,
) -> pl.DataFrame:
    g = _leer_fuente_ruta(leer_estudiantes_graduacion, ruta_archivo_graduacion(cfg, base, carpeta))
    p = _leer_fuente_ruta(leer_estudiantes_permanencia, ruta_archivo_permanencia(cfg, base, carpeta))
    if g.height == 0:
        return p
    if p.height == 0:
        return g

    por_id: dict[str, dict[str, str | None]] = {}
    for row in g.iter_rows(named=True):
        por_id[str(row["_id_key"])] = dict(row)
    for row in p.iter_rows(named=True):
        key = str(row["_id_key"])
        if key in por_id:
            _completar_valores(por_id[key], dict(row))
        else:
            por_id[key] = dict(row)
    return _df_desde_por_id(por_id)


def _es_encabezado_metas(vals: list[Any]) -> bool:
    norms = [_norm(v) for v in vals if not _es_vacio(v)]
    if not any(n in _ENCABEZADOS_META for n in norms):
        return False
    return any("meta" in n for n in norms)


def _mapa_columnas_meta(headers: list[Any]) -> dict[str, int]:
    headers_norm = [_norm(h) for h in headers]
    usados: set[int] = set()
    mapa: dict[str, int] = {}
    for clave, aliases in _ALIAS_META.items():
        idx = _indice_por_aliases(headers_norm, aliases, usados)
        if idx is not None:
            usados.add(idx)
            mapa[clave] = idx
    return mapa


def _periodo_en_prefijo(vals: list[Any], idx_programa: int) -> str | None:
    tope = idx_programa if idx_programa > 0 else min(4, len(vals))
    for i in range(tope):
        p = formatear_periodo_cod(vals[i])
        if p:
            return p
    return None


def _formatear_entero(val: Any) -> str:
    if _es_vacio(val):
        return ""
    texto = _texto(val).replace(",", ".")
    try:
        n = float(texto)
    except ValueError:
        return _texto(val)
    if n == int(n):
        return str(int(n))
    return f"{n:.1f}"


def _programa_meta_permitido(nombre: str) -> bool:
    n = _norm(nombre)
    if n == "institucional":
        return True
    return programa_es_permitido(nombre)


def _fila_meta_vacia(fila: dict[str, str]) -> bool:
    return not any(
        fila.get(k)
        for k in (
            "poblacion",
            "meta_num",
            "alcanzado_num",
            "meta_pct",
            "alcanzado_pct",
            "estado",
        )
    )


def _bloques_metas_hoja(ws) -> list[dict[str, Any]]:
    por_periodo: dict[str, dict[str, Any]] = {}
    orden: list[str] = []
    mapa: dict[str, int] | None = None
    periodo_actual: str | None = None
    en_proyeccion = False
    vacias = 0

    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if all(_es_vacio(v) for v in vals):
            if mapa is not None:
                vacias += 1
                if vacias >= 25:
                    break
            continue
        vacias = 0
        joined = " ".join(_norm(v) for v in vals if not _es_vacio(v))
        if "proyec" in joined and not _es_encabezado_metas(vals):
            palabras = [p for p in joined.split() if p]
            if (
                periodo_actual
                and periodo_actual in por_periodo
                and palabras
                and all(p.startswith("proyec") for p in palabras)
            ):
                por_periodo[periodo_actual]["proyeccion"] = True
            en_proyeccion = True
            continue
        if _es_encabezado_metas(vals):
            mapa = _mapa_columnas_meta(vals)
            continue
        if not mapa or "programa" not in mapa:
            continue
        idx_prog = mapa["programa"]
        if idx_prog >= len(vals):
            continue
        periodo_fila = _periodo_en_prefijo(vals, idx_prog)
        if periodo_fila:
            periodo_actual = periodo_fila
        programa = _texto(vals[idx_prog])
        if not programa or _norm(programa) in _FILAS_TITULO_META:
            continue
        if not periodo_actual or not _programa_meta_permitido(programa):
            continue

        def _cel(clave: str) -> Any:
            idx = mapa.get(clave)
            if idx is None or idx >= len(vals):
                return None
            return vals[idx]

        fila = {
            "programa": programa,
            "periodo_inicio": formatear_periodo_cod(_cel("periodo_inicio"))
            or _texto_campo(_cel("periodo_inicio"))
            or "",
            "poblacion": _formatear_entero(_cel("poblacion")),
            "meta_num": _formatear_entero(_cel("meta_num")),
            "alcanzado_num": _formatear_entero(_cel("alcanzado_num")),
            "faltante_num": _formatear_entero(_cel("faltante_num")),
            "meta_pct": _numero_pct(_cel("meta_pct")) or "",
            "alcanzado_pct": _numero_pct(_cel("alcanzado_pct")) or "",
            "faltante_pct": _numero_pct(_cel("faltante_pct")) or "",
            "estado": (_texto_campo(_cel("estado")) or "").upper(),
        }
        estado = fila["estado"]
        fila["cumple"] = True if estado == "CUMPLE" else False if estado == "NO CUMPLE" else None
        if _fila_meta_vacia(fila):
            continue
        bloque = por_periodo.get(periodo_actual)
        if bloque is None:
            bloque = {
                "periodo": periodo_actual,
                "proyeccion": en_proyeccion,
                "filas": [],
            }
            por_periodo[periodo_actual] = bloque
            orden.append(periodo_actual)
        elif en_proyeccion:
            bloque["proyeccion"] = True
        bloque["filas"].append(fila)
    return [por_periodo[p] for p in orden]


def _hoja_es_historico(nombre: str) -> bool:
    return "histor" in _norm(nombre)


def _numero_pct_chart(val: Any) -> float | None:
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = float(val)
        if -1.5 <= n <= 1.5:
            n *= 100
        return round(n, 2)
    texto = _texto(val).replace("%", "").replace(",", ".")
    if not texto:
        return None
    try:
        n = float(texto)
    except ValueError:
        return None
    if -1.5 <= n <= 1.5:
        n *= 100
    return round(n, 2)


def _bloques_historico_hoja(ws) -> list[dict[str, Any]]:
    bloques: list[dict[str, Any]] = []
    actual: dict[str, Any] | None = None
    cols: dict[str, int] | None = None
    vacias = 0

    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if all(_es_vacio(v) for v in vals):
            vacias += 1
            if vacias >= 20:
                break
            continue
        vacias = 0
        norms = [_norm(v) for v in vals]
        textos = [_texto(v) for v in vals]
        unidos = " ".join(n for n in norms if n)

        if "periodo" in unidos and "perman" in unidos and "gradu" in unidos:
            cols = {"periodo": None, "perm_meta": None, "perm_cum": None, "grad_meta": None, "grad_cum": None}
            idx_perm = next((i for i, n in enumerate(norms) if n == "permanencia"), None)
            idx_grad = next((i for i, n in enumerate(norms) if "gradu" in n), None)
            idx_per = next((i for i, n in enumerate(norms) if n == "periodo"), None)
            if idx_per is not None:
                cols["periodo"] = idx_per
            if idx_perm is not None:
                cols["perm_meta"] = idx_perm
                cols["perm_cum"] = idx_perm + 1
            if idx_grad is not None:
                cols["grad_meta"] = idx_grad
                cols["grad_cum"] = idx_grad + 1
            continue

        if cols and unidos.replace(" ", "") in {"metacumplimiento metacumplimiento", "meta cumplimiento"}:
            continue
        if cols and all(n in {"", "meta", "cumplimiento"} for n in norms if n):
            if any(n == "meta" for n in norms):
                continue

        no_vacios = [(i, t) for i, t in enumerate(textos) if t]
        if len(no_vacios) == 1 and not formatear_periodo_cod(no_vacios[0][1]):
            nombre = no_vacios[0][1].strip()
            if _norm(nombre) not in {"periodo", "permanencia", "graduacion", "meta", "cumplimiento"}:
                actual = {"programa": nombre, "filas": []}
                bloques.append(actual)
                continue

        if not cols or cols.get("periodo") is None:
            continue
        idx_p = cols["periodo"]
        if idx_p >= len(vals):
            continue
        periodo = formatear_periodo_cod(vals[idx_p]) or (
            _texto(vals[idx_p]) if re.match(r"^\d{4}\s*[-–]\s*[12]$", _texto(vals[idx_p])) else None
        )
        if not periodo:
            continue
        if actual is None:
            actual = {"programa": "Facultad", "filas": []}
            bloques.append(actual)

        def _cel_pct(clave: str) -> tuple[str, float | None]:
            i = cols.get(clave)
            if i is None or i >= len(vals):
                return "", None
            num = _numero_pct_chart(vals[i])
            texto = _numero_pct(vals[i]) or ""
            return texto, num

        p_meta_t, p_meta_n = _cel_pct("perm_meta")
        p_cum_t, p_cum_n = _cel_pct("perm_cum")
        g_meta_t, g_meta_n = _cel_pct("grad_meta")
        g_cum_t, g_cum_n = _cel_pct("grad_cum")
        actual["filas"].append(
            {
                "periodo": periodo,
                "permanencia_meta": p_meta_t,
                "permanencia_cumplimiento": p_cum_t,
                "graduacion_meta": g_meta_t,
                "graduacion_cumplimiento": g_cum_t,
                "permanencia_meta_n": p_meta_n,
                "permanencia_cumplimiento_n": p_cum_n,
                "graduacion_meta_n": g_meta_n,
                "graduacion_cumplimiento_n": g_cum_n,
            }
        )
    return [b for b in bloques if b.get("filas")]


def _pct_de_texto(val: Any) -> float | None:
    return _numero_pct_chart(val)


def _grafica_bloque_meta(tipo: str, bloque: dict[str, Any]) -> dict[str, Any] | None:
    filas = bloque.get("filas") or []
    if not filas:
        return None
    labels = [str(f.get("programa") or "") for f in filas]
    meta_pct = [_pct_de_texto(f.get("meta_pct")) for f in filas]
    alc_pct = [_pct_de_texto(f.get("alcanzado_pct")) for f in filas]
    usar_pct = any(v is not None for v in meta_pct + alc_pct)
    if usar_pct:
        datasets = [
            {"label": "Meta %", "data": meta_pct},
            {"label": "% alcanzado", "data": alc_pct},
        ]
        ylabel = "%"
    else:
        datasets = [
            {"label": "Meta #", "data": [_pct_de_texto(f.get("meta_num")) for f in filas]},
            {"label": "# alcanzado", "data": [_pct_de_texto(f.get("alcanzado_num")) for f in filas]},
        ]
        ylabel = "Estudiantes"
    periodo = bloque.get("periodo") or ""
    titulo = f"Metas de {tipo} · {periodo}".strip(" ·")
    return {
        "id": f"{tipo}:{periodo}",
        "etiqueta": titulo,
        "tipo": "bar",
        "titulo": titulo,
        "labels": labels,
        "datasets": datasets,
        "ylabel": ylabel,
    }


def _grafica_historico(bloque: dict[str, Any], clave: str) -> dict[str, Any] | None:
    filas = [
        f
        for f in (bloque.get("filas") or [])
        if f.get(f"{clave}_meta") or f.get(f"{clave}_cumplimiento")
    ]
    if not filas:
        return None
    programa = bloque.get("programa") or "Programa"
    nombre = "Permanencia" if clave == "permanencia" else "Graduación"
    return {
        "titulo": f"{nombre} · {programa}",
        "tipo": "line",
        "labels": [f.get("periodo") for f in filas],
        "datasets": [
            {"label": "Meta", "data": [f.get(f"{clave}_meta_n") for f in filas]},
            {"label": "Cumplimiento", "data": [f.get(f"{clave}_cumplimiento_n") for f in filas]},
        ],
        "ylabel": "%",
    }


def _grafica_historico_par(bloque: dict[str, Any]) -> dict[str, Any] | None:
    perm = _grafica_historico(bloque, "permanencia")
    grad = _grafica_historico(bloque, "graduacion")
    paneles = [p for p in (perm, grad) if p]
    if not paneles:
        return None
    programa = bloque.get("programa") or "Programa"
    return {
        "id": f"historico:{programa}",
        "etiqueta": f"Histórico · {programa}",
        "tipo": "historico",
        "titulo": f"Histórico · {programa}",
        "paneles": paneles,
    }


def _graficas_desde_metas(
    graduacion: list[dict[str, Any]],
    permanencia: list[dict[str, Any]],
    historico: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for b in graduacion:
        g = _grafica_bloque_meta("graduacion", b)
        if g:
            g["etiqueta"] = f"Graduación · {b.get('periodo')}"
            out.append(g)
    for b in permanencia:
        g = _grafica_bloque_meta("permanencia", b)
        if g:
            g["etiqueta"] = f"Permanencia · {b.get('periodo')}"
            out.append(g)
    for b in historico:
        g = _grafica_historico_par(b)
        if g:
            out.append(g)
    return out


def _hoja_es_proyeccion_graduacion(nombre: str) -> bool:
    n = _norm(nombre)
    return n in {"hoja1", "hoja 1"} or n.startswith("proyecc")


def leer_metas(ruta: Path) -> dict[str, Any]:
    vacio = {
        "disponible": False,
        "graduacion": [],
        "permanencia": [],
        "historico": [],
        "graficas": [],
    }
    if not ruta or not ruta.is_file():
        return vacio

    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        graduacion: list[dict[str, Any]] = []
        permanencia: list[dict[str, Any]] = []
        historico: list[dict[str, Any]] = []
        for nombre in wb.sheetnames:
            if _hoja_es_historico(nombre):
                historico.extend(_bloques_historico_hoja(wb[nombre]))
            elif _hoja_es_metas_permanencia(nombre):
                permanencia.extend(_bloques_metas_hoja(wb[nombre]))
            elif _hoja_es_metas_graduacion(nombre) or _hoja_es_proyeccion_graduacion(nombre):
                graduacion.extend(_bloques_metas_hoja(wb[nombre]))
    finally:
        wb.close()

    graficas = _graficas_desde_metas(graduacion, permanencia, historico)
    return {
        "disponible": bool(graduacion or permanencia or historico),
        "graduacion": graduacion,
        "permanencia": permanencia,
        "historico": historico,
        "graficas": graficas,
    }


def cargar_metas(
    cfg: dict,
    base: Path,
    carpeta: Path | None = None,
) -> dict[str, Any]:
    ruta = ruta_archivo_permanencia(cfg, base, carpeta)
    if ruta is None:
        return {
            "disponible": False,
            "graduacion": [],
            "permanencia": [],
            "historico": [],
            "graficas": [],
        }
    return leer_metas(ruta)
