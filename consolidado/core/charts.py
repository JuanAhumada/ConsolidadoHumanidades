"""Agregación de datos para gráficas (Chart.js en el frontend)."""

from __future__ import annotations

import re
from collections import Counter
from io import BytesIO
from typing import Any

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

TIPOS_GRAFICA = (
    "bar",
    "pie",
    "line",
)

_TIPO_POWERBI = {
    "bar": "Columnas agrupadas",
    "pie": "Gráfico circular",
    "line": "Gráfico de líneas",
}

_ALIAS_TIPO = {
    "bar_horizontal": "bar",
    "barras": "bar",
    "doughnut": "pie",
    "dona": "pie",
    "pastel": "pie",
    "torta": "pie",
    "linea": "line",
    "línea": "line",
}


def columnas_graficables(df: pl.DataFrame) -> list[str]:
    return list(df.columns)


def _partir_categorias(texto: str) -> list[str]:
    """Si el valor trae varios ítems unidos con «|», cada uno cuenta aparte."""
    partes = [p.strip() for p in str(texto).replace("||", "|").split("|") if p.strip()]
    return partes or [str(texto).strip()]


def preparar_datos_grafica(
    df: pl.DataFrame,
    *,
    columna: str,
    tipo: str = "bar",
    top: int = 25,
) -> dict[str, Any]:
    """
    Devuelve labels/valores listos para Chart.js.
    """
    tipo = _ALIAS_TIPO.get((tipo or "bar").strip().lower(), (tipo or "bar").strip().lower())
    if tipo not in TIPOS_GRAFICA:
        raise ValueError(f"Tipo no soportado: {tipo}. Use: {', '.join(TIPOS_GRAFICA)}")
    if columna not in df.columns:
        raise ValueError(f"Columna «{columna}» no encontrada.")

    valores: list[str] = []
    for v in df.get_column(columna).to_list():
        if v is None:
            continue
        texto = str(v).strip()
        if not texto or texto.lower() in {"none", "null", "nan"}:
            continue
        valores.extend(_partir_categorias(texto))

    if not valores:
        raise ValueError(f"La columna «{columna}» no tiene datos para graficar.")

    cont = Counter(valores)
    mas = cont.most_common(max(1, min(top, 50)))
    labels = [k for k, _ in mas]
    data = [n for _, n in mas]

    return {
        "columna": columna,
        "tipo": tipo,
        "chart_type": tipo,
        "horizontal": False,
        "labels": labels,
        "values": data,
        "total_filas": len(valores),
        "categorias": len(labels),
    }


def _nombre_hoja_excel(indice: int, columna: str, usados: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]+", "", f"G{indice}_{columna}".strip()) or f"Grafica{indice}"
    base = base[:31]
    nombre = base
    n = 2
    while nombre.lower() in usados:
        suf = f"_{n}"
        nombre = f"{base[: 31 - len(suf)]}{suf}"
        n += 1
    usados.add(nombre.lower())
    return nombre


def excel_powerbi_desde_graficas(series: list[dict[str, Any]]) -> bytes:
    """
    Excel tabular listo para Power BI: una hoja-tabla por gráfica
    (categoría + conteo) y una hoja de índice con el visual sugerido.
    """
    if not series:
        raise ValueError("No hay gráficas para exportar.")

    wb = Workbook()
    indice = wb.active
    indice.title = "Indice"
    cabecera = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="0C6B63")
    indice.append(["Hoja", "Columna", "Visual sugerido en Power BI", "Categorias", "Valores"])
    for cell in indice[1]:
        cell.font = cabecera
        cell.fill = fondo

    usados: set[str] = {"indice"}
    for i, item in enumerate(series, start=1):
        columna = str(item.get("columna") or f"Grafica {i}")
        tipo = str(item.get("tipo") or "bar")
        labels = list(item.get("labels") or [])
        values = list(item.get("values") or [])
        hoja_nombre = _nombre_hoja_excel(i, columna, usados)
        ws = wb.create_sheet(hoja_nombre)
        ws.append([columna, "Conteo"])
        for lab, val in zip(labels, values):
            ws.append([lab, val])
        ultima = max(len(labels) + 1, 2)
        for cell in ws[1]:
            cell.font = cabecera
            cell.fill = fondo
        for row in ws.iter_rows(min_row=2, max_row=ultima, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "#,##0"
        tabla = Table(displayName=f"Grafica{i}", ref=f"A1:B{ultima}")
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabla)
        ws.column_dimensions["A"].width = max(18, min(42, max((len(str(x)) for x in [columna, *labels]), default=12) + 2))
        ws.column_dimensions["B"].width = 14
        ws.auto_filter.ref = f"A1:B{ultima}"
        ws.freeze_panes = "A2"
        indice.append(
            [
                hoja_nombre,
                columna,
                _TIPO_POWERBI.get(tipo, "Columnas agrupadas"),
                len(labels),
                int(sum(float(v) for v in values if v is not None)),
            ]
        )

    uso = wb.create_sheet("Como_importar", 1)
    uso["A1"] = "Cómo llevar estas gráficas a Power BI"
    uso["A1"].font = Font(bold=True, size=14, color="0C6B63")
    lineas = [
        "",
        "Opción A — Pegar (una gráfica):",
        "1. En la web pulse «Copiar datos».",
        "2. En Power BI Desktop: Inicio → Introducir datos.",
        "3. Clic en la primera celda y Ctrl+V. Cargue.",
        "4. Inserte el visual indicado en la hoja Índice.",
        "",
        "Opción B — Excel (varias gráficas):",
        "1. Power BI Desktop: Inicio → Obtener datos → Excel.",
        "2. Elija este archivo. Verá una tabla por gráfica (Grafica1, Grafica2…).",
        "3. Seleccione las tablas y pulse Cargar.",
        "4. Arrastre la columna de categoría y Conteo al visual.",
        "",
        "Las imágenes PNG sirven para Insertar → Imagen, pero no son visuales interactivos.",
    ]
    for i, texto in enumerate(lineas, start=2):
        uso[f"A{i}"] = texto
        uso[f"A{i}"].alignment = Alignment(wrap_text=True)
    uso.column_dimensions["A"].width = 88
    for col in indice.columns:
        letra = get_column_letter(col[0].column)
        indice.column_dimensions[letra].width = 28
    indice.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
