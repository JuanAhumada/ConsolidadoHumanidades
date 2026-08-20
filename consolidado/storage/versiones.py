"""
Importar/exportar Excel de versión y semilla inicial.

asegurar_semilla_si_vacia corre al arrancar la web si no hay versiones.
Datos antiguos usa carpeta_fuentes=datos/historico para no tocar entrada.
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from consolidado.config.settings import (
    ETIQUETAS_EXPORT_COLUMNAS,
    cargar_config,
    construir_columnas_salida,
)
from consolidado.core.constants import HOJA_LISTADO, max_materias_en_dataframe
from consolidado.paths import PROJECT_ROOT
from consolidado.storage.db import (
    cargar_dataframe_version,
    contar_versiones,
    guardar_version,
    inicializar_db,
    nombre_excel_version,
    obtener_version,
    periodo_desde_fecha,
)


def _max_materias_desde_encabezados(headers: list[str | None]) -> int:
    max_n = 1
    for h in headers:
        if not h:
            continue
        m = re.match(r"^Materia\s+(\d+)$", str(h).strip(), re.IGNORECASE)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return max_n


def leer_dataframe_desde_excel_consolidado(
    ruta: Path,
    *,
    cfg: dict | None = None,
    base: Path | None = None,
) -> tuple[pl.DataFrame, int]:
    """
    Lee un Excel generado por la app (fila 1 = grupos, fila 2 = columnas).
    Mapea etiquetas de exportación a nombres canónicos por posición.
    """
    base = base or PROJECT_ROOT
    cfg = cfg or cargar_config(base)
    nombre_hoja = cfg.get("salida", {}).get("hoja", HOJA_LISTADO)

    wb = load_workbook(ruta, data_only=True, read_only=True)
    try:
        ws = wb[nombre_hoja] if nombre_hoja in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            next(rows)  # fila de grupos
            headers_raw = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"El Excel no tiene encabezados válidos: {ruta}") from exc

        headers = [str(h).strip() if h is not None else "" for h in headers_raw]
        while headers and headers[-1] == "":
            headers.pop()
        num_materias = _max_materias_desde_encabezados(headers)
        columnas = construir_columnas_salida(cfg, num_materias)

        # Si el Excel tiene exactamente el mismo número de columnas, usar orden canónico.
        # Si no, intentar mapear por etiqueta de exportación.
        reverse_etiquetas: dict[str, str] = {
            v: k for k, v in ETIQUETAS_EXPORT_COLUMNAS.items()
        }
        nombres: list[str] = []
        usados_canon: set[str] = set()
        if len(headers) == len(columnas):
            nombres = list(columnas)
        else:
            for h in headers:
                if not h:
                    continue
                if h in columnas and h not in usados_canon:
                    nombres.append(h)
                    usados_canon.add(h)
                elif h in reverse_etiquetas and reverse_etiquetas[h] not in usados_canon:
                    canon = reverse_etiquetas[h]
                    nombres.append(canon)
                    usados_canon.add(canon)
                elif h not in usados_canon:
                    # Desambiguar duplicados tipo Priorizado / Repitiendo
                    if h == "Priorizado" and "Ptje Priorizado" not in usados_canon:
                        nombres.append("Ptje Priorizado")
                        usados_canon.add("Ptje Priorizado")
                    elif h == "Repitiendo" and "Ptje Repitiendo" not in usados_canon:
                        nombres.append("Ptje Repitiendo")
                        usados_canon.add("Ptje Repitiendo")
                    else:
                        nombres.append(h)
                        usados_canon.add(h)

        datos: list[dict] = []
        for row in rows:
            valores = list(row[: len(nombres)])
            if all(v is None or str(v).strip() == "" for v in valores):
                continue
            fila = {nombres[i]: valores[i] if i < len(valores) else None for i in range(len(nombres))}
            datos.append(fila)
    finally:
        wb.close()

    if not datos:
        df = pl.DataFrame({c: [] for c in (nombres or columnas)})
    else:
        # infer_schema_length=None evita fallos por bool/null mixtos en columnas
        df = pl.from_dicts(datos, infer_schema_length=None)
        for c in columnas:
            if c not in df.columns:
                df = df.with_columns(pl.lit(None).alias(c))
        df = df.select([c for c in columnas if c in df.columns])

    return df, num_materias


def sembrar_version_inicial(
    base: Path | None = None,
    *,
    forzar: bool = False,
    fecha_version: date | None = None,
    periodo: str | None = None,
    excel_origen: Path | None = None,
) -> dict | None:
    """
    Crea el registro inicial 2026-1 (mayo) con los datos del último consolidado Excel.
    No hace nada si ya hay versiones, salvo forzar=True.
    """
    base = base or PROJECT_ROOT
    inicializar_db(base)
    if not forzar and contar_versiones(base) > 0:
        return None

    cfg = cargar_config(base)
    if excel_origen is None:
        rel = cfg.get("salida", {}).get("ruta", "salida/estudiantes_consolidado.xlsx")
        excel_origen = base / rel
    if not excel_origen.is_file():
        # Buscar cualquier consolidado en salida/
        candidatos = sorted(
            (base / "salida").glob("estudiantes_consolidado*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if not candidatos:
            return None
        excel_origen = candidatos[0]

    fecha_version = fecha_version or date(2026, 5, 10)
    periodo = periodo or "2026-1"

    df, num_materias = leer_dataframe_desde_excel_consolidado(
        excel_origen, cfg=cfg, base=base
    )
    if df.height == 0:
        return None

    # Copiar/renombrar Excel de referencia con el esquema de versión
    carpeta_salida = (base / "salida")
    carpeta_salida.mkdir(parents=True, exist_ok=True)
    nombre = nombre_excel_version(periodo, fecha_version)
    destino_excel = carpeta_salida / nombre

    # Regenerar Excel con formato de la app desde el dataframe leído
    from consolidado.core.export import guardar_excel_consolidado

    destino_excel = guardar_excel_consolidado(
        df, destino_excel, cfg=cfg, num_materias=num_materias
    )

    return guardar_version(
        df,
        base=base,
        fecha_version=fecha_version,
        periodo=periodo,
        num_materias=num_materias,
        ruta_excel=destino_excel,
        notas="Registro inicial importado del consolidado existente (periodo 2026-1, mayo).",
    )


def importar_excel_como_version(
    excel_origen: Path,
    *,
    fecha_version: date,
    notas: str | None = None,
    base: Path | None = None,
) -> dict:
    """Crea un snapshot SQL a partir de un Excel consolidado ya generado."""
    from consolidado.core.export import guardar_excel_consolidado
    from consolidado.core.pipeline import resolver_destino_versionado

    base = base or PROJECT_ROOT
    inicializar_db(base)
    cfg = cargar_config(base)
    origen = Path(excel_origen)
    if not origen.is_file():
        raise ValueError(f"No se encontró el Excel: {origen}")

    df, num_materias = leer_dataframe_desde_excel_consolidado(
        origen, cfg=cfg, base=base
    )
    if df.height == 0:
        raise ValueError("El Excel no contiene estudiantes.")

    destino, periodo, fecha_v = resolver_destino_versionado(
        cfg, base, fecha_version=fecha_version
    )
    destino = guardar_excel_consolidado(
        df, destino, cfg=cfg, num_materias=num_materias
    )
    texto_notas = (notas or "").strip() or (
        f"Excel importado · periodo {periodo} · {fecha_v.isoformat()}"
    )
    return guardar_version(
        df,
        base=base,
        fecha_version=fecha_v,
        periodo=periodo,
        num_materias=num_materias,
        ruta_excel=destino,
        notas=texto_notas,
    )


def ruta_excel_version(version_id: int, base: Path | None = None) -> Path | None:
    """Ruta del Excel asociado si el archivo todavía existe."""
    base = base or PROJECT_ROOT
    meta = obtener_version(version_id, base)
    if not meta or not meta.get("ruta_excel"):
        return None
    ruta = Path(str(meta["ruta_excel"]))
    if not ruta.is_absolute():
        ruta = base / ruta
    return ruta if ruta.is_file() else None


def exportar_excel_version(version_id: int, base: Path | None = None) -> Path:
    """Regenera el Excel desde el snapshot SQL (no crea una versión nueva)."""
    from consolidado.core.constants import max_materias_en_dataframe
    from consolidado.core.export import guardar_excel_consolidado
    from consolidado.core.repetidas import _cargar_materias_repetidas_cfg

    base = base or PROJECT_ROOT
    meta = obtener_version(version_id, base)
    if meta is None:
        raise ValueError(f"No existe la versión id={version_id}")
    df = cargar_dataframe_version(version_id, base)
    cfg = cargar_config(base)
    num_materias = meta.get("num_materias") or max_materias_en_dataframe(df)
    materias_repetidas = _cargar_materias_repetidas_cfg(cfg, base)
    fecha = date.fromisoformat(str(meta["fecha_version"]))
    periodo = str(meta.get("periodo") or periodo_desde_fecha(fecha))
    carpeta = base / "salida"
    carpeta.mkdir(parents=True, exist_ok=True)
    destino = carpeta / nombre_excel_version(periodo, fecha, sufijo_hora="export")
    return guardar_excel_consolidado(
        df,
        destino,
        cfg=cfg,
        num_materias=num_materias,
        materias_repetidas=materias_repetidas,
    )


def asegurar_excel_version(version_id: int, base: Path | None = None) -> Path:
    """Devuelve el Excel de la versión; si falta, lo regenera desde SQL."""
    existente = ruta_excel_version(version_id, base)
    if existente is not None:
        return existente
    return exportar_excel_version(version_id, base)


def asegurar_semilla_si_vacia(base: Path | None = None) -> dict | None:
    """Invocado al arrancar la app: si la BD está vacía, intenta sembrar 2026-1."""
    base = base or PROJECT_ROOT
    inicializar_db(base)
    if contar_versiones(base) > 0:
        return None
    return sembrar_version_inicial(base)
